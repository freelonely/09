from openpyxl import load_workbook

class ExcelMethod():
    def __init__(self,logger):
        self.logger = logger
    def readExcel(self,filename):
        self.filename=filename
        excel_list=[]
        try:
            self.wb=load_workbook(filename)
            self.sheet = self.wb.worksheets[0]
            self.max_column = self.sheet.max_column
            for row in self.sheet.rows:
                son_list = []
                for cell in row:
                    son_list.append(cell.value)
                excel_list.append(son_list)
        except:
            self.logger.error((filename,'加载失败'))
        else:
            return excel_list

    def saveExcel(self,row,text):
        try:
            self.sheet.cell(row,self.max_column,text)
            self.wb.save(self.filename)
        except:
            self.logger.error((self.filename,'保存失败'))
